# בדיקות M17 — ייצוא לידים לאקסל + כלל החיפוש המשותף
"""M17 strict tests: the leads Excel export + the shared keyword-match rule.

No DB, no session, no network: `build_leads_workbook` takes plain dicts (the
`list_leads` shape) and `lead_matches` is pure, so both are testable directly.
What we pin down here:

  * the workbook OPENS and its header row is exactly the fixed Hebrew columns
    followed by the union of answer keys in first-seen order;
  * a FILE answer writes the file name and carries an absolute hyperlink to the
    session-gated /api/leads/files/{id} route;
  * a customer answer of "=cmd|..." is NEUTRALIZED (leading apostrophe) so Excel
    cannot execute it;
  * an empty result set still produces a valid header-only workbook;
  * `lead_matches` follows the documented rule, empty query included.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app.services.leads.export import (
    _FIXED_HEADERS,
    build_export_filename,
    build_leads_workbook,
)
from app.services.leads.search import lead_matches

BASE = "https://app.example.com"


def _lead(**over):
    lead = {
        "id": "11111111-1111-1111-1111-111111111111",
        "lead_name": "שיפוצים",
        "phone": "+972500000000",
        "contact_name": "דנה כהן",
        "answers": {"מה השם שלך?": "דנה", "כמה חדרים?": "3"},
        "status": "new",
        "close_reason": None,
        "outcome_note": None,
        "last_step_index": 2,
        "is_test": False,
        "conversation_id": "c1",
        "started_at": "2026-07-01T08:30:00+00:00",
        "last_activity_at": "2026-07-02T09:00:00+00:00",
        "submitted_at": None,
        "feed_seen_at": None,
    }
    lead.update(over)
    return lead


def _open(leads):
    data = build_leads_workbook(leads, base_url=BASE, sheet_title="שיפוצים")
    return load_workbook(BytesIO(data))


# --- the workbook ------------------------------------------------------------


def test_headers_are_fixed_columns_then_answer_keys_in_first_seen_order():
    leads = [
        _lead(answers={"שאלה א": "1", "שאלה ב": "2"}),
        _lead(answers={"שאלה ב": "9", "שאלה ג": "3"}),
    ]
    ws = _open(leads).active
    header = [c.value for c in ws[1]]
    assert header[: len(_FIXED_HEADERS)] == list(_FIXED_HEADERS)
    assert header[len(_FIXED_HEADERS):] == ["שאלה א", "שאלה ב", "שאלה ג"]
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref is not None
    assert ws.max_row == 3  # header + 2 leads


def test_file_answer_writes_name_and_absolute_hyperlink():
    fid = "aaaaaaaa-bbbb-cccc-dddd-eeeeeeeeeeee"
    leads = [
        _lead(
            answers={
                "צרפו תמונה": {
                    "file_id": fid,
                    "mime_type": "image/png",
                    "name": "תוכנית.png",
                }
            }
        )
    ]
    ws = _open(leads).active
    cell = ws.cell(row=2, column=len(_FIXED_HEADERS) + 1)
    assert cell.value == "תוכנית.png"
    assert cell.hyperlink is not None
    assert cell.hyperlink.target == f"{BASE}/api/leads/files/{fid}"


def test_multiple_files_share_one_cell_and_link_the_first():
    a, b = "11111111-0000-0000-0000-000000000001", "22222222-0000-0000-0000-000000000002"
    leads = [
        _lead(
            answers={
                "מסמכים": [
                    {"file_id": a, "mime_type": "application/pdf", "name": "חוזה.pdf"},
                    {"file_id": b, "mime_type": "application/pdf", "name": "נספח.pdf"},
                ]
            }
        )
    ]
    ws = _open(leads).active
    cell = ws.cell(row=2, column=len(_FIXED_HEADERS) + 1)
    assert cell.value == "חוזה.pdf\nנספח.pdf"
    assert cell.hyperlink.target == f"{BASE}/api/leads/files/{a}"


def test_formula_injection_is_neutralized():
    hostile = '=cmd|\' /C calc\'!A0'
    leads = [_lead(answers={"הערה": hostile}, contact_name="-2+3", outcome_note="@SUM(1)")]
    ws = _open(leads).active
    assert ws.cell(row=2, column=len(_FIXED_HEADERS) + 1).value == "'" + hostile
    assert ws.cell(row=2, column=1).value == "'-2+3"       # שם
    assert ws.cell(row=2, column=9).value == "'@SUM(1)"    # סיכום הטיפול


def test_status_and_dates_render():
    leads = [_lead(status="deal", close_reason="completed")]
    ws = _open(leads).active
    assert ws.cell(row=2, column=4).value == "בוצעה עסקה"
    assert ws.cell(row=2, column=5).value == "ליד הושלם"
    started = ws.cell(row=2, column=6)
    assert hasattr(started.value, "year")  # a real datetime, not an ISO string
    assert started.number_format == "DD/MM/YYYY HH:MM"


def test_empty_result_still_produces_a_valid_header_only_workbook():
    ws = _open([]).active
    assert [c.value for c in ws[1]] == list(_FIXED_HEADERS)
    assert ws.max_row == 1


def test_filename_weaves_in_the_active_filters():
    assert build_export_filename().startswith("כל הפניות — ")
    assert build_export_filename().endswith(".xlsx")
    assert "למסלול שיפוצים" in build_export_filename(flow="שיפוצים")
    assert "(בוצעה עסקה)" in build_export_filename(status="deal")
    # 'all' adds nothing; a hostile flow name cannot inject a header/newline.
    assert "(" not in build_export_filename(status="all")
    assert "\n" not in build_export_filename(flow='a\r\nb/c"')
    assert "/" not in build_export_filename(flow='a\r\nb/c"')


# --- the shared match rule ---------------------------------------------------


def test_empty_query_matches_everything():
    lead = _lead()
    for q in ("", "   ", "\t", None):
        assert lead_matches(lead, q) is True


def test_matches_plain_fields_case_insensitively():
    lead = _lead(contact_name="Dana Cohen", outcome_note="Called Back")
    assert lead_matches(lead, "dana")
    assert lead_matches(lead, "  COHEN ")
    assert lead_matches(lead, "called")
    assert lead_matches(lead, "97250")         # phone substring (plain, no normalizing)
    assert lead_matches(lead, "שיפוצ")          # lead_name
    assert not lead_matches(lead, "zzz")


def test_matches_answer_keys_values_and_file_names():
    lead = _lead(
        answers={
            "כמה חדרים?": "ארבעה",
            "קובץ": {"file_id": "x", "mime_type": "image/png", "name": "Blueprint.PNG"},
            "מסמכים": [
                {"file_id": "y", "mime_type": "application/pdf", "name": "חוזה.pdf"},
                "טקסט חופשי",
            ],
        }
    )
    assert lead_matches(lead, "חדרים")        # a key
    assert lead_matches(lead, "ארבעה")        # a string value
    assert lead_matches(lead, "blueprint")    # a file name, case-insensitive
    assert lead_matches(lead, "חוזה")         # a file name inside a list
    assert lead_matches(lead, "חופשי")        # a string inside a list
    assert not lead_matches(lead, "file_id")  # plumbing is NOT searchable
    assert not lead_matches(lead, "image/png")


def test_none_fields_are_skipped_not_stringified():
    lead = _lead(contact_name=None, outcome_note=None, answers={})
    assert not lead_matches(lead, "none")
