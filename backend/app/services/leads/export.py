# מחברת הלידים — ייצוא לאקסל: בניית חוברת עבודה מהלידים של בעל העסק
"""Excel (.xlsx) export of the owner's leads — the workbook builder (M17).

`GET /api/leads/export` hands us the ALREADY-FILTERED, already-decrypted rows
(the exact same `leads.list_leads` output the dashboard renders) and gets back
the bytes of a formatted workbook. Keeping the building here keeps the API route
thin and makes the whole thing unit-testable without a DB or a session.

Design notes worth knowing before you edit:

  * WRITE-ONLY MODE. `openpyxl.Workbook(write_only=True)` streams rows straight
    to the temp XML instead of holding a cell grid in RAM. The prod box has 1 GB
    and the export asks `list_leads` for up to EXPORT_MAX_ROWS (50k) — far above
    the list view's 500 — so a flat footprint is what makes that safe. This is
    the cap rising that the note used to hedge against. Write-only mode means: rows only via
    `ws.append([...])`, cells must be built as `WriteOnlyCell`, and anything
    sheet-level (freeze panes, autofilter, column widths) must be set BEFORE the
    first append.

  * FORMULA INJECTION. Every answer, name and note in here is CUSTOMER-supplied
    text. Excel treats a cell starting with `=`, `+`, `-` or `@` as a formula,
    which is a real code-execution vector when the owner opens the file (the
    classic CSV-injection bug). `_safe_text` prefixes those with an apostrophe.
    Do not bypass it for any customer-derived string.

  * PII. Names, phones and answers are exactly the plaintext this module exists
    to write into a file for the owner. NOTHING here logs a value — the caller
    logs a row count and the non-PII filters only.

Column order: the fixed Hebrew columns first (see `_FIXED_HEADERS`), then ONE
column per distinct answer key. The answer columns are ordered by FIRST-SEEN
across the exported rows (rows come back newest-first, so the newest lead's
question order leads). That is stable for a given result set and needs no
knowledge of the bot config — which may have changed since an old lead ran.
"""

from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:  # pragma: no cover - depends on tzdata being present in the image
    from zoneinfo import ZoneInfo

    _LOCAL_TZ: Any = ZoneInfo("Asia/Jerusalem")
except Exception:  # pragma: no cover - fall back to UTC rather than crash
    _LOCAL_TZ = timezone.utc


# --- Hebrew label maps (MIRRORED FROM THE FRONTEND) --------------------------
# Source of truth for the wording:
#   status       → frontend/src/components/dashboard/LeadCard.tsx  (STATUS_META)
#   close_reason → frontend/src/dashboard/closeReason.ts           (CLOSE_REASON_META)
# The export must read like the dashboard, so if a label changes there, change it
# here too. Unknown values fall through as the raw code rather than an empty cell.
STATUS_LABELS: dict[str, str] = {
    "new": "חדש",
    "in_progress": "פתוח",
    "abandoned": "ננטש",
    "deal": "בוצעה עסקה",
    "closed": "ליד סגור",
}

CLOSE_REASON_LABELS: dict[str, str] = {
    "completed": "ליד הושלם",
    "abandoned": "ליד ננטש",
    "answered": "מענה הושלם",
}


# The FILTER vocabulary is slightly wider than the stored statuses: the leads
# list also accepts 'all' (no filter) and the synthetic 'open' (= new +
# in_progress). Only these are woven into the download's file name.
_FILTER_STATUS_LABELS: dict[str, str] = {
    **STATUS_LABELS,
    "open": "בטיפול",
    "all": "",
}


def status_label(value: str | None) -> str:
    """The dashboard's Hebrew word for a stored status ('' when unset)."""
    if not value:
        return ""
    return STATUS_LABELS.get(value, value)


def close_reason_label(value: str | None) -> str:
    """The dashboard's Hebrew word for a stored close_reason ('' when unset)."""
    if not value:
        return ""
    return CLOSE_REASON_LABELS.get(value, value)


# --- sheet shape -------------------------------------------------------------

# The leading columns, in order. Everything after these is one-per-answer-key.
_FIXED_HEADERS: tuple[str, ...] = (
    "שם",
    "טלפון",
    "מסלול",
    "סטטוס",
    "סיבת סגירה",
    "תאריך התחלה",
    "תאריך השלמה",
    "פעילות אחרונה",
    "סיכום הטיפול",
)

# Which fixed columns hold a date (used for the number format + column width).
_DATE_COLUMN_INDEXES = (6, 7, 8)  # 1-based: התחלה / השלמה / פעילות אחרונה

_DATE_FORMAT = "DD/MM/YYYY HH:MM"

_HEADER_FILL = PatternFill("solid", fgColor="E8F0E4")  # light leaf-ish tint
_HEADER_FONT = Font(bold=True, color="1F2937")
_LINK_FONT = Font(color="0563C1", underline="single")

# readingOrder=2 == "right-to-left" in the OOXML alignment spec (0=context,
# 1=LTR, 2=RTL). Every cell in this workbook is Hebrew, so we state the
# direction explicitly rather than leaving it to the viewer to infer — see the
# column loop in `build_leads_workbook` for why inference is not good enough.
# Tab colours, one per flow, cycled. Deliberately a MUTED set: every entry sits
# in the same narrow band of lightness and saturation (desaturated mid-tones), so
# a workbook with six questionnaires reads as one considered palette rather than
# a rainbow. They differ in hue only — enough to tell tabs apart at a glance,
# never enough for one tab to shout louder than the rest. Keep any addition in
# the same band.
_TAB_COLORS: tuple[str, ...] = (
    "A8BFA3",  # sage
    "A9BCD0",  # dusty blue
    "D0B9A0",  # sand
    "BFAEC4",  # muted mauve
    "A6C1BC",  # soft teal
    "C9BFA0",  # khaki
    "C7ADA8",  # dusty rose
    "B3B9C9",  # slate
)

_RTL = Alignment(readingOrder=2)
_RTL_HEADER = Alignment(readingOrder=2, vertical="center", wrap_text=True)
_RTL_WRAP_TOP = Alignment(readingOrder=2, wrap_text=True, vertical="top")

# Column widths in Excel character units. The fixed columns get purpose-built
# widths; answer columns share one comfortable default.
_FIXED_WIDTHS = (22, 16, 20, 14, 14, 18, 18, 18, 36)
_ANSWER_WIDTH = 32

# Excel's own hard cap on the characters in one cell.
_MAX_CELL_CHARS = 32767

# The characters Excel will read as the start of a formula.
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _safe_text(value: Any) -> str:
    """Customer text → a string Excel will always treat as TEXT, never a formula.

    Prefixes an apostrophe when the value starts with `=`, `+`, `-` or `@` (the
    CSV/formula-injection guard) and truncates to Excel's per-cell limit. Also
    strips the control characters that would make the XML invalid.
    """
    if value is None:
        return ""
    text = value if isinstance(value, str) else str(value)
    # Tabs/newlines are legal in a cell; the other C0 controls are not.
    text = "".join(ch for ch in text if ch in "\t\n\r" or ord(ch) >= 32)
    if text[:1] in _FORMULA_PREFIXES:
        text = "'" + text
    return text[:_MAX_CELL_CHARS]


def _parse_dt(value: Any) -> datetime | None:
    """ISO-8601 string (as `list_leads` returns) → a naive local datetime.

    We write REAL datetimes (with a number format) rather than ISO text so Excel
    can sort and filter them. Excel has no timezone concept, so the value is
    converted to Israel local time and the tzinfo is dropped — matching what the
    owner sees in the dashboard, which renders in the browser's local zone.
    """
    if isinstance(value, datetime):
        parsed = value
    elif isinstance(value, str) and value:
        try:
            parsed = datetime.fromisoformat(value)
        except ValueError:
            return None
    else:
        return None

    if parsed.tzinfo is None:
        # A naive value out of the DB is UTC by convention (timestamptz).
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(_LOCAL_TZ).replace(tzinfo=None)


# --- answer rendering --------------------------------------------------------


def _file_answers(value: Any) -> list[dict]:
    """Every file answer carried by one value ([] for a plain string).

    Mirrors `fileAnswers` in frontend/src/components/dashboard/AnswerValue.tsx:
    a file answer is a dict with a string `file_id`; a list may hold several.
    """
    if isinstance(value, dict) and isinstance(value.get("file_id"), str):
        return [value]
    if isinstance(value, list):
        return [
            item
            for item in value
            if isinstance(item, dict) and isinstance(item.get("file_id"), str)
        ]
    return []


def _file_href(base_url: str, file_id: str) -> str:
    """The owner-only, session-gated download URL for one stored file.

    Absolute on purpose: a workbook is opened OUTSIDE the browser tab, so a
    relative `/api/...` link (what the UI uses) would be meaningless. There are
    still no presigned URLs — the link points at the same gated route, and the
    owner must be logged in for it to resolve.
    """
    return f"{base_url}/api/leads/files/{file_id}"


def _answer_cell(ws: Any, value: Any, base_url: str) -> Any:
    """Build ONE answer cell: plain text, or file name(s) with a hyperlink.

    MULTI-FILE RENDERING (the documented choice): each file's NAME goes on its
    own LINE inside the single cell (wrapped text), and the cell's hyperlink
    points at the FIRST file. Excel allows only one hyperlink per cell, so the
    alternative — one cell per file — would break the "one column per question"
    shape. The remaining files are still identifiable by name and reachable from
    the dashboard, and the common case (exactly one file) is a perfect link.
    """
    files = _file_answers(value)
    if files:
        names = [
            _safe_text(f.get("name") or "").strip() or "קובץ" for f in files
        ]
        cell = WriteOnlyCell(ws, value="\n".join(names))
        cell.hyperlink = _file_href(base_url, str(files[0]["file_id"]))
        cell.font = _LINK_FONT
        cell.alignment = _RTL_WRAP_TOP
        return cell

    if isinstance(value, str):
        return _safe_text(value)
    if value is None:
        return ""
    # Numbers/bools can be written as-is; anything else becomes readable text.
    if isinstance(value, (int, float, bool)):
        return value
    return _safe_text(str(value))


# --- the workbook ------------------------------------------------------------


def _answer_keys(leads: list[dict]) -> list[str]:
    """The union of answer keys, in FIRST-SEEN order across `leads`.

    `dict` preserves insertion order, so a plain "add if new" walk gives a stable
    ordering for a given result set without sorting the questions alphabetically
    (which would scramble a questionnaire's natural flow).
    """
    keys: dict[str, None] = {}
    for lead in leads:
        answers = lead.get("answers")
        if isinstance(answers, dict):
            for key in answers:
                if isinstance(key, str):
                    keys.setdefault(key, None)
    return list(keys)


def _sheet_title(raw: str | None, taken: set[str]) -> str:
    """A legal, UNIQUE Excel sheet title derived from a flow name.

    Excel caps titles at 31 chars and rejects  : \\ / ? * [ ]  — a flow name is
    owner-authored free text, so both rules have to be enforced here rather than
    trusted. Collisions are real once names are truncated to 31 chars (two long
    flows can sanitize to the same string), and openpyxl would silently rename
    the second one; we disambiguate with a numeric suffix instead so a tab is
    never mysteriously called "לידים1".
    """
    title = (raw or "לידים").strip() or "לידים"
    title = "".join(ch for ch in title if ch not in ':\\/?*[]')[:31] or "לידים"
    if title not in taken:
        taken.add(title)
        return title
    for n in range(2, 100):
        suffix = f" ({n})"
        candidate = title[: 31 - len(suffix)] + suffix
        if candidate not in taken:
            taken.add(candidate)
            return candidate
    taken.add(title)
    return title


def _group_by_flow(leads: list[dict], fallback: str) -> list[tuple[str, list[dict]]]:
    """Split leads into (flow label, rows), BIGGEST flow first.

    Rows arrive newest-first and already filtered. Grouping is by `lead_name`,
    the same value the dashboard's "מסלול" filter uses, so a tab corresponds
    exactly to one filter choice. Ordering by descending row count puts the
    owner's busiest questionnaire in the first tab (the one Excel opens on);
    ties keep first-seen order, so the result is deterministic for a given
    result set.
    """
    groups: dict[str, list[dict]] = {}
    for lead in leads:
        name = lead.get("lead_name")
        label = name.strip() if isinstance(name, str) and name.strip() else fallback
        groups.setdefault(label, []).append(lead)
    return sorted(groups.items(), key=lambda kv: -len(kv[1]))


def build_leads_workbook(
    leads: list[dict],
    *,
    base_url: str,
    sheet_title: str | None = None,
) -> bytes:
    """Render the exported leads as .xlsx bytes — ONE TAB PER FLOW.

    `leads` are rows exactly as `leads.list_leads` returns them (decrypted, and
    already filtered by the caller). `base_url` is `settings.public_base_url`
    with any trailing slash removed — it prefixes the file hyperlinks.

    Leads are grouped by `lead_name` into one worksheet each, biggest first, and
    every tab gets a colour off `_TAB_COLORS`. Splitting this way also fixes a
    real readability problem: answer columns are computed PER SHEET, so a
    "ביטוח רכב" tab shows only the car questions instead of a wide, mostly-empty
    grid holding every question from every questionnaire.

    An EMPTY list is valid and produces a real workbook holding just the header
    row (titled from `sheet_title`); the endpoint never 404s on "no results".
    """
    base_url = (base_url or "").rstrip("/")
    fallback = (sheet_title or "לידים").strip() or "לידים"

    wb = Workbook(write_only=True)
    taken: set[str] = set()
    groups = _group_by_flow(leads, fallback) or [(fallback, [])]
    for index, (label, rows) in enumerate(groups):
        _build_flow_sheet(
            wb,
            title=_sheet_title(label, taken),
            leads=rows,
            base_url=base_url,
            tab_color=_TAB_COLORS[index % len(_TAB_COLORS)],
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _build_flow_sheet(
    wb: Workbook,
    *,
    title: str,
    leads: list[dict],
    base_url: str,
    tab_color: str,
) -> None:
    """Append ONE worksheet holding a single flow's leads."""
    keys = _answer_keys(leads)
    headers = list(_FIXED_HEADERS) + keys
    ws = wb.create_sheet(title=title)
    ws.sheet_properties.tabColor = tab_color

    # Sheet-level setup MUST happen before the first append in write-only mode.
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(leads) + 1, 1)}"
    for idx in range(1, len(headers) + 1):
        width = (
            _FIXED_WIDTHS[idx - 1]
            if idx <= len(_FIXED_WIDTHS)
            else _ANSWER_WIDTH
        )
        col = ws.column_dimensions[get_column_letter(idx)]
        col.width = width
        # Force RTL reading order at the COLUMN level. `rightToLeft` below only
        # flips the sheet layout (column A on the right); it says nothing about
        # how text inside a cell is laid out. Without this, each cell falls back
        # to readingOrder=0 ("context"), which leaves the direction up to the
        # viewer's bidi implementation — and a viewer that resolves it as LTR
        # renders every Hebrew string backwards. Setting it on the column is
        # what makes plain string cells inherit it: in write-only mode the bulk
        # of cells are appended as bare `str` and carry no style of their own.
        col.alignment = _RTL
    # Hebrew workbook: lay the sheet out right-to-left like the dashboard.
    ws.sheet_view.rightToLeft = True

    header_cells = []
    for name in headers:
        cell = WriteOnlyCell(ws, value=_safe_text(name))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _RTL_HEADER
        header_cells.append(cell)
    ws.append(header_cells)

    for lead in leads:
        answers = lead.get("answers")
        answers = answers if isinstance(answers, dict) else {}

        row: list[Any] = [
            _safe_text(lead.get("contact_name")),
            _safe_text(lead.get("phone")),
            _safe_text(lead.get("lead_name")),
            _safe_text(status_label(lead.get("status"))),
            _safe_text(close_reason_label(lead.get("close_reason"))),
        ]
        for field in ("started_at", "submitted_at", "last_activity_at"):
            parsed = _parse_dt(lead.get(field))
            if parsed is None:
                row.append("")
            else:
                cell = WriteOnlyCell(ws, value=parsed)
                cell.number_format = _DATE_FORMAT
                row.append(cell)
        row.append(_safe_text(lead.get("outcome_note")))

        for key in keys:
            row.append(_answer_cell(ws, answers.get(key), base_url))

        ws.append(row)


# --- the download file name --------------------------------------------------


def build_export_filename(
    *,
    flow: str | None = None,
    status: str | None = None,
    today: datetime | None = None,
) -> str:
    """A human, Hebrew title for the downloaded file, built from the filters.

        no filters        →  "כל הפניות — 2026-07-20.xlsx"
        flow selected     →  "כל הפניות למסלול <flow> — 2026-07-20.xlsx"
        status selected   →  "... (<status label>) — 2026-07-20.xlsx"

    The flow name is customer/owner-supplied, so the result goes through the
    same hygiene as an uploaded file name (`file_storage.sanitize_filename`):
    no path separators, no quotes, no CR/LF that could forge a
    `Content-Disposition` header. The caller still emits BOTH the ASCII fallback
    and the RFC 5987 `filename*` form so the Hebrew survives.
    """
    # Imported here (not at module top) purely to keep this module importable
    # without the storage stack; file_storage itself has no hard boto3 import.
    from app.services import file_storage

    stamp = (today or datetime.now(_LOCAL_TZ)).strftime("%Y-%m-%d")

    title = "כל הפניות"
    if flow:
        title += f" למסלול {flow}"
    label = _FILTER_STATUS_LABELS.get(status or "all", "")
    if label:
        title += f" ({label})"

    return file_storage.sanitize_filename(
        f"{title} — {stamp}.xlsx", fallback=f"leads-{stamp}.xlsx"
    )
